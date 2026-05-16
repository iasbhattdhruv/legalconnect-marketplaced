import os
import random
import string
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.contrib.auth.models import User

logger = logging.getLogger(__name__)

EXCEL_FILE = os.path.join(settings.BASE_DIR, 'users.xlsx')
EXPECTED_HEADERS = [
    'username',
    'email',
    'password',
    'user_type',
    'gender',
    'birthdate',
    'profession',
    'photo',
    'signature',
    'created_at'
]


def _normalize_header(header):
    if header is None:
        return ''
    return str(header).strip().lower().replace(' ', '_')


def _clean_users_excel():
    if not os.path.exists(EXCEL_FILE):
        return

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    clean_rows = []

    for row in sheet.iter_rows(values_only=True):
        if any(cell not in (None, '') for cell in row):
            clean_rows.append(row)

    if len(clean_rows) != sheet.max_row:
        workbook = Workbook()
        ws = workbook.active
        for row in clean_rows:
            ws.append(row)
        workbook.save(EXCEL_FILE)


def ensure_users_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(EXPECTED_HEADERS)
        wb.save(EXCEL_FILE)
        return EXCEL_FILE

    _clean_users_excel()
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    headers = [str(cell.value).strip().lower().replace(' ', '_') if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        for header in missing:
            sheet.cell(row=1, column=len(headers) + 1).value = header
            headers.append(header)
        workbook.save(EXCEL_FILE)
    return EXCEL_FILE


def append_user_to_excel(user, password=None):
    try:
        ensure_users_excel()
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        headers = [str(cell.value).strip().lower().replace(' ', '_') if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        birthdate_value = ''
        if hasattr(user, 'profile') and user.profile.birthdate:
            if hasattr(user.profile.birthdate, 'isoformat'):
                birthdate_value = user.profile.birthdate.isoformat()
            else:
                birthdate_value = str(user.profile.birthdate)

        row_data = {
            'username': user.username,
            'email': user.email,
            'password': '',
            'user_type': user.profile.user_type if hasattr(user, 'profile') else '',
            'gender': user.profile.gender if hasattr(user, 'profile') else '',
            'birthdate': birthdate_value,
            'profession': user.profile.profession if hasattr(user, 'profile') else '',
            'photo': user.profile.photo.url if hasattr(user, 'profile') and user.profile.photo else '',
            'signature': user.profile.signature.url if hasattr(user, 'profile') and user.profile.signature else '',
            'created_at': datetime.now().isoformat(),
        }
        row = [row_data.get(header, '') for header in headers]
        sheet.append(row)
        workbook.save(EXCEL_FILE)
    except OSError as exc:
        logger.warning("Could not write users.xlsx; user was saved in database. Error: %s", exc)


def _generate_password(length=12):
    characters = string.ascii_letters + string.digits + '!@#$%^&*()'
    return ''.join(random.choice(characters) for _ in range(length))


def sync_users_from_excel():
    ensure_users_excel()
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip().lower().replace(' ', '_') if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {header: idx for idx, header in enumerate(headers) if header}

    created = 0
    skipped = 0
    errors = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {header: row[idx] if idx < len(row) else None for header, idx in header_map.items()}
        username = str(row_data.get('username', '')).strip() if row_data.get('username') else ''
        email = str(row_data.get('email', '')).strip() if row_data.get('email') else ''
        if not username or not email:
            skipped += 1
            continue

        if User.objects.filter(username=username).exists() or User.objects.filter(email=email).exists():
            skipped += 1
            continue

        password = str(row_data.get('password')).strip() if row_data.get('password') else _generate_password()
        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.save()
            profile = user.profile
            profile.user_type = str(row_data.get('user_type', '')).strip() or 'client'
            profile.gender = str(row_data.get('gender', '')).strip()
            birthdate = row_data.get('birthdate')
            if isinstance(birthdate, str) and birthdate:
                try:
                    birthdate = datetime.fromisoformat(birthdate).date()
                except ValueError:
                    birthdate = None
            profile.birthdate = birthdate
            profile.profession = str(row_data.get('profession', '')).strip()
            profile.save()
            created += 1
        except Exception as exc:
            errors.append(str(exc))

    return {
        'created': created,
        'skipped': skipped,
        'errors': errors,
    }
