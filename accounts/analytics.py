import os
from collections import Counter
from datetime import datetime, timedelta
from django.conf import settings
from django.db.models import Count, Q, Sum
from .models import Appointment, Invoice, Profile
from .ml_features import operational_insights

# Keep the Excel functions for backward compatibility if needed
DATA_FILE = os.path.join(settings.BASE_DIR, 'data', 'legal_booking_dataset.xlsx')
EXPECTED_HEADERS = [
    'booking_id',
    'username',
    'legal_category',
    'service_type',
    'lawyer_id',
    'lawyer_name',
    'lawyer_specialization',
    'appointment_date',
    'appointment_time',
    'booking_status',
    'booking_channel',
    'created_at'
]


def _normalize_row(row, header_map):
    return {
        'booking_id': row[header_map['booking_id']],
        'username': row[header_map['username']],
        'legal_category': row[header_map['legal_category']],
        'service_type': row[header_map['service_type']],
        'lawyer_id': row[header_map['lawyer_id']],
        'lawyer_name': row[header_map['lawyer_name']],
        'lawyer_specialization': row[header_map['lawyer_specialization']],
        'appointment_date': row[header_map['appointment_date']],
        'appointment_time': row[header_map['appointment_time']],
        'booking_status': row[header_map['booking_status']],
        'booking_channel': row[header_map['booking_channel']],
        'created_at': row[header_map['created_at']],
    }


def load_booking_dataset():
    # This function remains for any legacy use, but we'll use database queries now
    if not os.path.exists(DATA_FILE):
        return []

    try:
        from openpyxl import load_workbook
        workbook = load_workbook(DATA_FILE, data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows or len(rows) < 2:
            return []

        headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        missing_keys = [key for key in EXPECTED_HEADERS if key not in header_map]
        if missing_keys:
            return []

        entries = []
        for values in rows[1:]:
            if not values or not values[0]:
                continue
            entry = _normalize_row(values, header_map)
            entries.append(entry)

        return entries
    except Exception:
        return []


def _format_top_items(counter, limit=3):
    return [
        {'label': label or 'Unknown', 'count': count}
        for label, count in counter.most_common(limit)
    ]


def summarize_booking_data():
    # Use real database data instead of Excel
    appointments = Appointment.objects.all()

    if not appointments:
        return {
            'total_bookings': 0,
            'top_category': '—',
            'top_category_count': 0,
            'top_service': '—',
            'top_service_count': 0,
            'peak_time': '—',
            'peak_time_count': 0,
            'top_lawyers': [],
            'status_counts': {},
            'channel_counts': {},
            'confirmed_rate': '0%',
            'latest_bookings': [],
            'daily_trend': [],
            'next_three_day_forecast': '0',
            'top_specializations': [],
            'last_data_entry': None,
            'acceptance_rate': '0%',
            'cancellation_rate': '0%',
            'total_revenue': 0,
            'pending_revenue': 0,
            'lawyer_utilization': [],
            'high_risk_appointments': [],
            'case_category_mix': [],
        }

    total = appointments.count()

    lawyer_profiles = {p.user_id: p for p in Profile.objects.filter(user_type='lawyer')}
    category_counts = Counter()
    daily_counts = Counter()

    for apt in appointments:
        lawyer_profile = lawyer_profiles.get(apt.lawyer_id)
        specialization = lawyer_profile.specialization if lawyer_profile and lawyer_profile.specialization else 'General'
        category_counts[specialization] += 1
        if apt.appointment_date:
            daily_counts[apt.appointment_date] += 1

    service_counts = Counter(['Legal Consultation' for _ in appointments])

    lawyer_booking_counts = appointments.values('lawyer_id', 'lawyer__username').annotate(count=Count('lawyer')).order_by('-count')
    lawyer_counts = Counter()
    for item in lawyer_booking_counts:
        lawyer_id = item['lawyer_id']
        lawyer_name = item['lawyer__username']
        lawyer_profile = lawyer_profiles.get(lawyer_id)
        display_name = f"{lawyer_profile.user.first_name} {lawyer_profile.user.last_name}" if lawyer_profile else lawyer_name
        lawyer_counts[display_name] = item['count']

    status_counts = Counter([apt.status for apt in appointments])
    channel_counts = Counter(['Web' for _ in appointments])
    time_counts = Counter([str(apt.appointment_time) for apt in appointments if apt.appointment_time])

    top_category, top_category_count = category_counts.most_common(1)[0] if category_counts else ('—', 0)
    top_service, top_service_count = service_counts.most_common(1)[0] if service_counts else ('—', 0)
    peak_time, peak_time_count = time_counts.most_common(1)[0] if time_counts else ('—', 0)
    top_lawyers = _format_top_items(lawyer_counts, limit=5)
    top_specializations = _format_top_items(category_counts, limit=4)

    accepted_count = status_counts.get('accepted', 0) + status_counts.get('Accepted', 0)
    confirmed_rate = f"{round(accepted_count / total * 100)}%" if total else '0%'
    cancelled_count = status_counts.get('cancelled', 0) + status_counts.get('Cancelled', 0)
    cancellation_rate = f"{round(cancelled_count / total * 100)}%" if total else '0%'

    trend_items = sorted(daily_counts.items())
    daily_trend = [
        {'date': item[0].isoformat(), 'count': item[1]}
        for item in trend_items
    ]

    next_three_day_forecast = '0'
    if len(trend_items) >= 3:
        try:
            import pandas as pd
            df = pd.DataFrame(daily_trend)
            df['date'] = pd.to_datetime(df['date'])
            df['day_number'] = df['date'].map(pd.Timestamp.toordinal)
            df['count'] = df['count'].astype(float)
            X = df[['day_number']]
            y = df['count']
            slope = ((X['day_number'] - X['day_number'].mean()) * (y - y.mean())).sum() / ((X['day_number'] - X['day_number'].mean()) ** 2).sum()
            intercept = y.mean() - slope * X['day_number'].mean()
            future_days = [trend_items[-1][0].toordinal() + i for i in range(1, 4)]
            forecast_values = [max(0, round(slope * day + intercept)) for day in future_days]
            next_three_day_forecast = str(sum(forecast_values))
        except Exception:
            next_three_day_forecast = str(round(sum(count for _, count in trend_items[-3:]) / min(3, len(trend_items))))
    else:
        next_three_day_forecast = str(sum(count for _, count in trend_items[-3:]))

    latest_appointments = appointments.order_by('-created_at')[:3]
    latest_bookings = []
    for apt in latest_appointments:
        lawyer_profile = lawyer_profiles.get(apt.lawyer_id)
        lawyer_name = f"{lawyer_profile.user.first_name} {lawyer_profile.user.last_name}" if lawyer_profile else apt.lawyer.username
        latest_bookings.append({
            'booking_id': apt.id,
            'username': apt.client.username,
            'lawyer_name': lawyer_name,
            'appointment_date': apt.appointment_date,
            'appointment_time': apt.appointment_time,
            'booking_status': apt.status,
            'created_at': apt.created_at,
        })

    last_data_entry = latest_appointments[0].created_at if latest_appointments else None
    revenue_total = Invoice.objects.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    pending_revenue = Invoice.objects.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0

    utilization = []
    for profile in Profile.objects.filter(user_type='lawyer').select_related('user'):
        lawyer_total = appointments.filter(lawyer=profile.user).count()
        lawyer_pending = appointments.filter(lawyer=profile.user, status='pending').count()
        lawyer_accepted = appointments.filter(lawyer=profile.user, status='accepted').count()
        utilization.append({
            'label': profile.user.get_full_name() or profile.user.username,
            'specialization': profile.specialization or 'General',
            'total': lawyer_total,
            'pending': lawyer_pending,
            'accepted': lawyer_accepted,
            'load_score': min(100, lawyer_pending * 18 + lawyer_accepted * 7),
        })
    utilization.sort(key=lambda item: item['load_score'], reverse=True)
    ops = operational_insights()

    return {
        'total_bookings': total,
        'top_category': top_category,
        'top_category_count': top_category_count,
        'top_service': top_service,
        'top_service_count': top_service_count,
        'peak_time': peak_time,
        'peak_time_count': peak_time_count,
        'top_lawyers': top_lawyers,
        'status_counts': dict(status_counts),
        'channel_counts': dict(channel_counts),
        'confirmed_rate': confirmed_rate,
        'latest_bookings': latest_bookings,
        'daily_trend': daily_trend,
        'next_three_day_forecast': next_three_day_forecast,
        'top_specializations': top_specializations,
        'last_data_entry': last_data_entry,
        'acceptance_rate': confirmed_rate,
        'cancellation_rate': cancellation_rate,
        'total_revenue': revenue_total,
        'pending_revenue': pending_revenue,
        'lawyer_utilization': utilization[:5],
        'high_risk_appointments': ops['high_risk_appointments'],
        'case_category_mix': ops['case_category_mix'],
    }
